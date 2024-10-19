import datetime
import os
import openpyxl


def get_folder_path():
    current_directory = os.getcwd()
    fold_path = os.path.abspath(os.path.join(current_directory, 'excel_docs'))
    if not os.path.exists(fold_path):
        os.makedirs(fold_path)

    return fold_path


folder_path = get_folder_path()


def get_file_list():  # получить названия всех файлов excel в папке
    excel_files = [file for file in os.listdir(folder_path) if file.endswith('.xlsx') or file.endswith('.XLSX')]
    print(excel_files)
    return excel_files


def value_validation(w_data):
    if type(w_data) is datetime.datetime:
        return w_data.strftime('%d.%m.%Y')
    elif type(w_data) is datetime.time:
        return w_data.strftime('%H:%M')
    else:
        return w_data


def get_worker_data(file_name):  # получить информацию о работнике
    worker_data = {}
    wb = openpyxl.load_workbook(os.path.join(folder_path, file_name))
    sheet = wb.worksheets[1]
    worker_data['tab_num'] = sheet['A2'].value
    worker_data['month_start'] = value_validation(sheet['B2'].value)
    worker_data['month_end'] = value_validation(sheet['C2'].value)
    worker_data['post'] = sheet['D2'].value
    worker_data['fio'] = sheet['E2'].value
    worker_data['year'] = sheet['F2'].value
    worker_data['month'] = sheet['G2'].value
    worker_data['period'] = sheet['G2'].value
    wb.close()

    return worker_data


def get_amount_of_dates(file_name):
    wb = openpyxl.load_workbook(os.path.join(folder_path, file_name))
    sheet = wb.worksheets[1]
    amount_of_dates = 0
    i = 3
    while True:
        val = sheet[f'A{i}'].value
        if val is None:
            break
        else:
            amount_of_dates += 1
        i += 1

    return amount_of_dates


def get_grv_dates(file_name):
    i = 3
    dates = []
    wb = openpyxl.load_workbook(os.path.join(folder_path, file_name))
    sheet = wb.worksheets[1]
    amount = get_amount_of_dates(file_name) + i
    print('AMOUNT:', amount)
    while i < amount:
        # date = sheet[f'A{i}'].value.strftime('%d.%m.%Y')
        date = value_validation(sheet[f'A{i}'].value)
        work_s = value_validation(sheet[f'B{i}'].value)
        if work_s == 'МП':
            work_s = 'МП'
            work_e = ' '
            rest_s = ' '
            rest_e = ' '
        else:
            work_s = value_validation(sheet[f'B{i}'].value)
            work_e = value_validation(sheet[f'C{i}'].value)
            rest_s = value_validation(sheet[f'D{i}'].value)
            if rest_s == '':
                rest_e = "None"
            else:
                rest_s = value_validation(sheet[f'D{i}'].value)
                rest_e = value_validation(sheet[f'E{i}'].value)

            if rest_s != 'Без ОП':
                pass
            else:
                rest_s = 'Без_ОП'

        date_time_str = str(date) + ' ' + str(work_s) + ' ' + str(work_e) + ' ' + str(rest_s) + ' ' + str(rest_e)
        print('date_time_str', date_time_str)
        dates.append(date_time_str)
        i += 1
    wb.close()

    return dates

